from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from typing import Optional
import csv
import io
from datetime import datetime

from models.candidate import Candidate, CandidateStatus
from models.job import Job

router = APIRouter()


async def get_candidates_with_jobs(job_id: Optional[str], status: Optional[str]):
    query = Candidate.find()
    if job_id:
        query = query.find(Candidate.job_id == job_id)
    if status and status in ("shortlisted", "rejected", "pending"):
        query = query.find(Candidate.status == CandidateStatus(status))
    candidates = await query.sort("-overall_score").to_list()

    # Fetch all jobs for title lookup
    all_jobs = await Job.find_all().to_list()
    job_map = {str(j.id): j.title for j in all_jobs}
    return candidates, job_map


@router.get("/csv")
async def export_csv(
    job_id: Optional[str] = Query(None),
    status: Optional[str] = Query("shortlisted"),
):
    """Export candidates as CSV."""
    candidates, job_map = await get_candidates_with_jobs(job_id, status)

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Rank", "Name", "Email", "Phone", "Job Title",
        "Overall Score (%)", "Skill Match (%)", "Experience Match (%)", "Semantic Match (%)",
        "Skills", "Experience", "Education",
        "Skill Gaps", "AI Summary", "Status", "Uploaded At"
    ])

    for rank, c in enumerate(candidates, 1):
        writer.writerow([
            rank,
            c.name,
            c.email,
            c.phone,
            job_map.get(c.job_id, ""),
            c.overall_score,
            c.skill_match,
            c.experience_match,
            c.semantic_match,
            ", ".join(c.skills),
            c.experience,
            c.education,
            ", ".join(c.skill_gap),
            c.ai_summary,
            c.status,
            c.uploaded_at.strftime("%Y-%m-%d %H:%M"),
        ])

    output.seek(0)
    filename = f"candidates_{status or 'all'}_{datetime.utcnow().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/excel")
async def export_excel(
    job_id: Optional[str] = Query(None),
    status: Optional[str] = Query("shortlisted"),
):
    """Export candidates as Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    candidates, job_map = await get_candidates_with_jobs(job_id, status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"

    headers = [
        "Rank", "Name", "Email", "Job Title", "Score (%)",
        "Skill Match (%)", "Exp Match (%)", "Skills", "Experience",
        "Skill Gaps", "Status"
    ]
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for rank, c in enumerate(candidates, 1):
        row = [
            rank, c.name, c.email,
            job_map.get(c.job_id, ""),
            c.overall_score, c.skill_match, c.experience_match,
            ", ".join(c.skills[:8]),
            c.experience,
            ", ".join(c.skill_gap[:5]),
            c.status,
        ]
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"candidates_{status or 'all'}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
